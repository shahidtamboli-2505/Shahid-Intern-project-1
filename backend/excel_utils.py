from __future__ import annotations

from typing import List, Dict, Any
import re
import json

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# -----------------------------
# Helpers
# -----------------------------
def _norm(s: Any) -> str:
    s = "" if s is None else str(s)
    return re.sub(r"\s+", " ", s.strip())


def _pick(row: Dict[str, Any], *keys: str, default: Any = "") -> Any:
    for k in keys:
        if k in row:
            val = row.get(k)
            if val not in (None, "", "null", "N/A", "na", "None", "NULL"):
                return val
    return default


def _yes_no(v: Any) -> str:
    if isinstance(v, bool):
        return "Yes" if v else "No"
    s = _norm(v).lower()
    return "Yes" if s in {"yes", "true", "1", "y"} else "No"


def _to_number_or_blank(x: Any) -> Any:
    try:
        s = str(x).strip()
        if not s or s.lower() in {"nan", "none", "null"}:
            return ""
        f = float(s.replace(",", ""))
        return int(f) if f.is_integer() else f
    except Exception:
        return ""


def _safe_json_load(x: Any) -> Any:
    if x is None:
        return None
    if isinstance(x, (dict, list)):
        return x
    if isinstance(x, str):
        s = x.strip()
        if not s:
            return None
        try:
            return json.loads(s)
        except Exception:
            return None
    return None


# -----------------------------
# FINAL EXCEL SCHEMA (LOCKED)
# -----------------------------
FINAL_COLS: List[str] = [
    "Company Name",
    "Industry",
    "Google Rating",
    "Rating Count",
    "Has Website",
    "Website URL",
    "Contact Phone",
    "Contact Email",
    "Address",
    "Place ID",
    "Source Name",
    "Source URL",
    "Leadership Found",
    "Name 1",
    "Designation 1",
    "Name 2",
    "Designation 2",
    "Name 3",
    "Designation 3",
    "Name 4",
    "Designation 4",
    "Name 5",
    "Designation 5",
]


# Buckets order (must match Case-2 structure)
BUCKETS_ORDER = [
    "Executive Leadership",
    "Technology / Operations",
    "Finance / Administration",
    "Business Development / Growth",
    "Marketing / Branding",
]


def _flatten_case2_management_to_names(case2_management: Any) -> Dict[str, str]:
    """
    Convert bucket dict -> Name 1..5, Designation 1..5 (strict order).
    Accepts dict or json-string.

    ✅ FIXED:
    - If name exists but designation missing, still fill (designation blank allowed).
    - designation fallback from role.
    """
    out: Dict[str, str] = {}
    for i in range(1, 6):
        out[f"Name {i}"] = ""
        out[f"Designation {i}"] = ""

    mgmt = _safe_json_load(case2_management) or {}
    if not isinstance(mgmt, dict):
        return out

    idx = 1
    for bucket in BUCKETS_ORDER:
        if idx > 5:
            break

        v = mgmt.get(bucket) or {}
        if not isinstance(v, dict):
            continue

        nm = _norm(v.get("name", ""))
        dg = _norm(v.get("designation", "")) or _norm(v.get("role", ""))

        # ✅ name-only allowed
        if nm:
            out[f"Name {idx}"] = nm
            out[f"Designation {idx}"] = dg
            idx += 1

    return out


def _flatten_case2_leaders_legacy(case2_leaders: Any) -> Dict[str, str]:
    """
    Legacy support: case2_leaders = [{name, role}, ...]
    ✅ Also supports designation key.
    """
    out: Dict[str, str] = {}
    for i in range(1, 6):
        out[f"Name {i}"] = ""
        out[f"Designation {i}"] = ""

    leaders = _safe_json_load(case2_leaders) or []
    if not isinstance(leaders, list):
        return out

    for i in range(1, 6):
        idx = i - 1
        if idx < len(leaders) and isinstance(leaders[idx], dict):
            out[f"Name {i}"] = _norm(leaders[idx].get("name", ""))
            out[f"Designation {i}"] = _norm(
                leaders[idx].get("role", leaders[idx].get("designation", ""))
            )
    return out


# -----------------------------
# Row Builder
# -----------------------------
def _build_excel_row(row: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        "Company Name": _norm(_pick(row, "Company Name", "name", "company_name", default="Unknown")),
        "Industry": _norm(_pick(row, "Industry", "industry", "Category", "primaryType", default="Business")),
        "Google Rating": _to_number_or_blank(_pick(row, "Google Rating", "rating", "google_rating")),
        "Rating Count": _to_number_or_blank(
            _pick(row, "Rating Count", "Reviews", "userRatingCount", "rating_count", "google_rating_count")
        ),
        "Has Website": "Yes" if _pick(row, "Website URL", "website", "websiteUri") else _yes_no(
            _pick(row, "Has Website", "has_website")
        ),
        "Website URL": _norm(_pick(row, "Website URL", "website", "websiteUri", "website_url")),
        "Contact Phone": _norm(
            _pick(
                row,
                "Contact Phone",
                "nationalPhoneNumber",
                "internationalPhoneNumber",
                "phone",
                "contact_phone",
            )
        ),
        "Contact Email": _norm(_pick(row, "Contact Email", "email", "contact_email")),
        "Address": _norm(_pick(row, "Address", "formattedAddress", "address")),
        "Place ID": _norm(_pick(row, "Place ID", "google_place_id", "place_id", "id")),
        "Source Name": _norm(_pick(row, "Source Name", "source_name", default="Google Places")) or "Google Places",
        "Source URL": _norm(_pick(row, "Source URL", "googleMapsUri", "url", "source_url")),
    }

    # ---------------------------------------------------------
    # Leaders priority
    # 1) If row already has Name/Designation 1..5 -> use them
    # 2) else if case2_management exists -> flatten buckets
    # 3) else legacy case2_leaders list
    # ---------------------------------------------------------
    has_flat = False
    for i in range(1, 6):
        n = _norm(row.get(f"Name {i}", ""))
        d = _norm(row.get(f"Designation {i}", ""))
        if n or d:
            has_flat = True
            break

    if has_flat:
        for i in range(1, 6):
            out[f"Name {i}"] = _norm(row.get(f"Name {i}", ""))
            out[f"Designation {i}"] = _norm(row.get(f"Designation {i}", ""))
    else:
        flat_from_mgmt = _flatten_case2_management_to_names(row.get("case2_management"))
        if any(flat_from_mgmt.get(f"Name {i}") for i in range(1, 6)):
            for i in range(1, 6):
                out[f"Name {i}"] = flat_from_mgmt.get(f"Name {i}", "")
                out[f"Designation {i}"] = flat_from_mgmt.get(f"Designation {i}", "")
        else:
            flat_legacy = _flatten_case2_leaders_legacy(row.get("case2_leaders"))
            for i in range(1, 6):
                out[f"Name {i}"] = flat_legacy.get(f"Name {i}", "")
                out[f"Designation {i}"] = flat_legacy.get(f"Designation {i}", "")

    # Leadership Found:
    lf = _norm(row.get("Leadership Found", ""))
    if lf in {"Yes", "No"}:
        out["Leadership Found"] = lf
    else:
        # ✅ FIX: if ANY Name i exists -> Yes
        out["Leadership Found"] = "Yes" if any(_norm(out.get(f"Name {i}", "")) for i in range(1, 6)) else "No"

    return out


# -----------------------------
# Excel Writer
# -----------------------------
def write_case1_excel(rows: List[Dict[str, Any]], out_path: str) -> None:
    normalized = [_build_excel_row(dict(r)) for r in (rows or [])]
    df = pd.DataFrame(normalized)

    if df.empty:
        df = pd.DataFrame(columns=FINAL_COLS)
    else:
        for col in FINAL_COLS:
            if col not in df.columns:
                df[col] = ""
        df = df[FINAL_COLS]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Mining Results")
        ws = writer.book["Mining Results"]

        ws.freeze_panes = "A2"

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Header styling
        for i, col_name in enumerate(FINAL_COLS, start=1):
            cell = ws.cell(row=1, column=i)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Column widths
        wide_cols = {"Company Name", "Website URL", "Source URL", "Contact Email", "Address"}
        for i, col_name in enumerate(FINAL_COLS, start=1):
            letter = get_column_letter(i)
            if col_name in wide_cols:
                ws.column_dimensions[letter].width = 42
            elif col_name == "Place ID":
                ws.column_dimensions[letter].width = 26
            elif "Name" in col_name or "Designation" in col_name:
                ws.column_dimensions[letter].width = 30
            else:
                ws.column_dimensions[letter].width = 16

        # Borders + wrap + Leadership Found conditional color
        lf_col = FINAL_COLS.index("Leadership Found") + 1
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            lf_cell = ws.cell(row=r, column=lf_col)
            lf_cell.fill = green_fill if str(lf_cell.value).strip() == "Yes" else red_fill

    print(f"✅ Excel Generation Successful: {out_path}")
